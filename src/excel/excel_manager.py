import os
import pandas as pd
from datetime import datetime
from pathlib import Path
import re

class ExcelManager:
    def __init__(self, output_folder):
        self.output_folder = output_folder
        self.excel_file_path = os.path.join(output_folder, "Date_Persoane_OCR.xlsx")
        self.data_list = []
    
    def extract_data_from_txt(self, txt_file_path):
        """Extrage datele dintr-un fișier .txt care a fost generat de process.py"""
        try:
            with open(txt_file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            # Calculam calea relativa fata de folderul de output
            relative_path = os.path.relpath(txt_file_path, self.output_folder)
            # Eliminam extensia .txt din calea relativa
            relative_path_no_ext = os.path.splitext(relative_path)[0]
            folder_name = os.path.dirname(relative_path) if os.path.dirname(relative_path) else "Root"
            
            # Initializam dictionarul cu date
            data = {
                'Nume_Fisier': os.path.splitext(os.path.basename(txt_file_path))[0],
                'Cale_Fisier': relative_path_no_ext,
                'Folder_Sursa': folder_name,
                'Data_Procesare': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'Text_Complet': content.strip()
            }
            
            # Parsam continutul fișierului text conform structurii din process.py
            # Structura: nume\ninitiala_tatalui\nprenume\ncnp_total\nadresa\nphone\nemail\ndoiani
            lines = content.strip().split('\n')
            
            print(f"Debug: Procesare fișier {txt_file_path}")
            print(f"Debug: Linii gasite: {lines}")
            
            # Extragem datele conform structurii exacte din process.py
            data['Nume'] = lines[0].strip() if len(lines) > 0 else ''
            data['Initiala_Tatalui'] = lines[1].strip() if len(lines) > 1 else ''
            data['Prenume'] = lines[2].strip() if len(lines) > 2 else ''
            
            # CNP - ne asiguram ca este tratat ca string pentru a evita notatia știintifica
            cnp_raw = lines[3].strip() if len(lines) > 3 else ''
            # Curatam CNP-ul și ne asiguram ca este string
            data['CNP'] = str(cnp_raw).strip() if cnp_raw else ''
            
            data['Adresa'] = lines[4].strip() if len(lines) > 4 else ''
            data['2_Ani'] = lines[7].strip() if len(lines) > 7 else ''  # doiani este pe pozitia 7 (index)
            
            # Informatii suplimentare (incluse in Excel)
            # Telefon - pastram intotdeauna ca string pentru a conserva zero-urile de la inceput
            telefon_raw = lines[5].strip() if len(lines) > 5 else ''
            if telefon_raw:
                # Eliminam doar .0 de la sfârșit daca exista, dar pastram zero-urile de la inceput
                telefon_clean = str(telefon_raw).strip()
                if telefon_clean.endswith('.0'):
                    telefon_clean = telefon_clean[:-2]  # Eliminam doar .0 de la sfârșit
                data['Telefon'] = telefon_clean
            else:
                data['Telefon'] = ''
            
            data['Email'] = lines[6].strip() if len(lines) > 6 else ''
            
            # ANAF de care apartin - folosim folder_localitate_mic (ultimul folder din ierarhie)
            data['ANAF_Apartin'] = self._get_folder_localitate_mic(folder_name)
            
            print(f"Debug: Rezultat final -> Nume: '{data['Nume']}', Initiala: '{data['Initiala_Tatalui']}', Prenume: '{data['Prenume']}', CNP: '{data['CNP']}', Telefon: '{data['Telefon']}', Email: '{data['Email']}', 2 Ani: '{data['2_Ani']}', ANAF: '{data['ANAF_Apartin']}'")
            
            return data
            
        except Exception as e:
            print(f"Eroare la procesarea fișierului {txt_file_path}: {e}")
            return None
    
    def _get_folder_localitate_mic(self, folder_name):
        """Extrage folder_localitate_mic din calea folderului"""
        if not folder_name or folder_name == "Root":
            return "NEDETERMINAT"
        
        # Împartim calea in parti
        parts = folder_name.split(os.sep)
        
        # Ultima parte din cale este folder_localitate_mic
        if len(parts) >= 1:
            folder_localitate_mic = parts[-1].strip()
            return folder_localitate_mic if folder_localitate_mic else "NEDETERMINAT"
        
        return "NEDETERMINAT"
    
    def _separate_name_parts(self, nume_complet):
        """Separa numele complet in nume, initiala tatalui și prenume"""
        if not nume_complet:
            return {'nume': '', 'initiala_tatalui': '', 'prenume': ''}
            
        # Curatam textul
        nume_complet = nume_complet.strip()
        parts = nume_complet.split()
        
        print(f"Debug: Separare nume pentru: '{nume_complet}' -> {parts}")
        
        if len(parts) == 0:
            return {'nume': '', 'initiala_tatalui': '', 'prenume': ''}
        elif len(parts) == 1:
            # Doar un cuvânt - probabil numele
            result = {'nume': parts[0], 'initiala_tatalui': '', 'prenume': ''}
            print(f"Debug: Un singur cuvânt -> {result}")
            return result
        elif len(parts) == 2:
            # Doua cuvinte - verificam daca al doilea pare initiala
            if len(parts[1]) <= 2 and parts[1].isupper():
                # Al doilea pare initiala
                result = {'nume': parts[0], 'initiala_tatalui': parts[1], 'prenume': ''}
                print(f"Debug: Doua cuvinte (cu initiala) -> {result}")
                return result
            else:
                # Nume și prenume
                result = {'nume': parts[0], 'initiala_tatalui': '', 'prenume': parts[1]}
                print(f"Debug: Doua cuvinte (nume + prenume) -> {result}")
                return result
        elif len(parts) == 3:
            # Trei cuvinte - cautam initiala
            # Verificam daca al doilea cuvânt pare sa fie o initiala
            if len(parts[1]) <= 2 and (parts[1].isupper() or parts[1].endswith('.')):
                result = {'nume': parts[0], 'initiala_tatalui': parts[1], 'prenume': parts[2]}
                print(f"Debug: Trei cuvinte (cu initiala in mijloc) -> {result}")
                return result
            # Verificam daca al treilea cuvânt pare sa fie o initiala
            elif len(parts[2]) <= 2 and (parts[2].isupper() or parts[2].endswith('.')):
                result = {'nume': parts[0], 'initiala_tatalui': parts[2], 'prenume': parts[1]}
                print(f"Debug: Trei cuvinte (cu initiala la sfârșit) -> {result}")
                return result
            else:
                # Probabil nume compus sau prenume compus
                result = {'nume': parts[0], 'initiala_tatalui': '', 'prenume': ' '.join(parts[1:])}
                print(f"Debug: Trei cuvinte (fara initiala) -> {result}")
                return result
        else:
            # Mai mult de trei cuvinte
            # Cautam o initiala (un-doua caractere, de preferinta mari)
            for i in range(1, len(parts)):
                if len(parts[i]) <= 2 and (parts[i].isupper() or parts[i].endswith('.')):
                    result = {
                        'nume': ' '.join(parts[:i]),
                        'initiala_tatalui': parts[i],
                        'prenume': ' '.join(parts[i+1:])
                    }
                    print(f"Debug: Multe cuvinte (cu initiala la pozitia {i}) -> {result}")
                    return result
            
            # Daca nu gasim initiala, impartim in nume și prenume
            result = {
                'nume': parts[0],
                'initiala_tatalui': '',
                'prenume': ' '.join(parts[1:])
            }
            print(f"Debug: Multe cuvinte (fara initiala) -> {result}")
            return result
    
    def add_person_data(self, txt_file_path):
        """Adauga datele unei persoane in lista pentru Excel"""
        data = self.extract_data_from_txt(txt_file_path)
        if data:
            self.data_list.append(data)
            return True
        return False
    
    def create_excel_file(self):
        """Creeaza fișierul Excel cu toate datele"""
        try:
            if not self.data_list:
                print("Nu exista date pentru a crea fișierul Excel.")
                return False
            
            # Cream DataFrame-ul
            df = pd.DataFrame(self.data_list)
            
            # Ne asiguram ca CNP-ul și Telefonul sunt tratate ca string pentru a evita formatarea automata
            if 'CNP' in df.columns:
                df['CNP'] = df['CNP'].astype(str)
            if 'Telefon' in df.columns:
                df['Telefon'] = df['Telefon'].astype(str)
            
            # Add a new column 'Validitate_CNP' next to 'CNP'
            if 'CNP' in df.columns:
                df['Validitate_CNP'] = df['CNP'].apply(lambda cnp: 'DA' if self.validate_cnp(cnp)[0] else 'NU')
            
            # Reordonam coloanele pentru a respecta ordinea ceruta
            preferred_columns = [
                'Nume', 'Initiala_Tatalui', 'Prenume', 'CNP', 'Validitate_CNP', 'Adresa', 'ANAF_Apartin', 'Telefon', 'Email', '2_Ani'
            ]
            
            # Adaugam coloanele suplimentare la sfârșit pentru referinta
            additional_columns = [
                'Nume_Fisier', 'Folder_Sursa', 'Cale_Fisier', 'Data_Procesare', 'Text_Complet'
            ]
            
            # Reordonam coloanele existente
            existing_main_columns = [col for col in preferred_columns if col in df.columns]
            existing_additional_columns = [col for col in additional_columns if col in df.columns]
            other_columns = [col for col in df.columns if col not in preferred_columns + additional_columns]
            
            final_columns = existing_main_columns + existing_additional_columns + other_columns
            
            df = df[final_columns]
            
            # Salvam in Excel cu formatare
            with pd.ExcelWriter(self.excel_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Date_Persoane', index=False)
                
                # Obtinem worksheet-ul pentru formatare
                worksheet = writer.sheets['Date_Persoane']
                
                # Formatam coloanele CNP și Telefon ca text pentru a evita formatarea automata
                from openpyxl.styles import NamedStyle
                from openpyxl.utils import get_column_letter
                
                # Gasim coloanele CNP și Telefon
                columns_to_format = {'CNP': None, 'Telefon': None}
                for idx, col_name in enumerate(df.columns, 1):
                    if col_name in columns_to_format:
                        columns_to_format[col_name] = idx
                        
                print(f"Debug: Coloane de formatat gasite: {columns_to_format}")
                
                # Formatam coloanele gasite
                for col_name, col_index in columns_to_format.items():
                    if col_index:
                        col_letter = get_column_letter(col_index)
                        # Formatam intreaga coloana ca text (inclusiv header-ul)
                        for row in range(1, len(df) + 2):  # De la header (1) pâna la ultimul rând
                            cell = worksheet[f'{col_letter}{row}']
                            cell.number_format = '@'  # Format text
                            
                            # Pentru datele din rândurile de continut (nu header)
                            if row > 1 and cell.value is not None:
                                # Convertim totul la string, indiferent de continut
                                original_value = str(cell.value).strip()
                                
                                # Pentru CNP, daca este in notatie știintifica, il convertim
                                if col_name == 'CNP' and ('E+' in original_value or 'e+' in original_value or '.' in original_value):
                                    try:
                                        # Încercam sa convertim din notatie știintifica
                                        cell.value = str(int(float(original_value)))
                                    except:
                                        cell.value = original_value
                                # Pentru telefon, eliminam .0 daca exista, dar pastram zero-urile de la inceput  
                                elif col_name == 'Telefon' and original_value.endswith('.0'):
                                    # Eliminam doar .0 de la sfârșit, pastrând zero-urile de la inceput
                                    cell.value = original_value[:-2]
                                else:
                                    # Pentru toate celelalte, il pastram ca string
                                    cell.value = original_value
                
                # Ajustam latimea coloanelor
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Maxim 50 caractere
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Fișierul Excel a fost creat cu succes: {self.excel_file_path}")
            print(f"Au fost procesate {len(self.data_list)} persoane.")
            
            # Generam automat un raport de validare
            print("\n📋 Generez raport de validare...")
            try:
                validation_result, duplicate_result = self.generate_validation_report()
                
                # Exportam automat și in CSV pentru backup
                csv_path = self.export_to_csv()
                if csv_path:
                    print(f"✅ Backup CSV creat automat: {os.path.basename(csv_path)}")
                
                # Generam automat și raportul PDF
                pdf_path = self.export_to_pdf_report()
                if pdf_path:
                    print(f"✅ Raport PDF generat automat: {os.path.basename(pdf_path)}")
                    
            except Exception as e:
                print(f"⚠️  Eroare la generarea raportului de validare: {e}")
                import traceback
                traceback.print_exc()
            
            return True
            
        except Exception as e:
            print(f"Eroare la crearea fișierului Excel: {e}")
            return False
    
    def process_all_txt_files(self, folder_path):
        """Proceseaza toate fișierele .txt dintr-un folder și toate subfolderele sale"""
        txt_files_found = []
        
        # Cautam recursiv in toate folderele și subfolderele
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith('.txt'):
                    txt_path = os.path.join(root, file)
                    txt_files_found.append(txt_path)
        
        if not txt_files_found:
            print("Nu au fost gasite fișiere .txt in folderul specificat și subfolderele sale.")
            return False
        
        print(f"Au fost gasite {len(txt_files_found)} fișiere .txt in folderul și subfolderele sale:")
        for txt_file in txt_files_found:
            print(f"  - {txt_file}")
            self.add_person_data(txt_file)
        
        return self.create_excel_file()

    def _determine_anaf(self, adresa, folder_name):
        """Determina ANAF-ul pe baza adresei sau folder-ului"""
        if not adresa and folder_name == "Root":
            return "NEDETERMINAT"
        
        # Daca avem informatii din folder
        if folder_name and folder_name != "Root":
            # Încearca sa extragi informatii ANAF din numele folderului
            folder_lower = folder_name.lower()
            if "anaf" in folder_lower:
                return folder_name
        
        # Daca avem adresa, incearca sa determinisectorul/judetul
        if adresa:
            adresa_lower = adresa.lower()
            
            # București - determinare sector
            if "bucuresti" in adresa_lower or "sector" in adresa_lower:
                for i in range(1, 7):
                    if f"sector {i}" in adresa_lower or f"sectorul {i}" in adresa_lower:
                        return f"ANAF SECTOR {i}"
                return "ANAF BUCURESTI"
            
            # Judete comune
            judete_anaf = {
                'cluj': 'ANAF CLUJ',
                'timis': 'ANAF TIMIS',
                'constanta': 'ANAF CONSTANTA',
                'iasi': 'ANAF IASI',
                'brasov': 'ANAF BRASOV',
                'galati': 'ANAF GALATI',
                'dolj': 'ANAF DOLJ',
                'arad': 'ANAF ARAD',
                'sibiu': 'ANAF SIBIU',
                'bacau': 'ANAF BACAU',
                'prahova': 'ANAF PRAHOVA',
                'maramures': 'ANAF MARAMURES',
                'bihor': 'ANAF BIHOR',
                'mures': 'ANAF MURES',
                'suceava': 'ANAF SUCEAVA'
            }
            
            for judet, anaf in judete_anaf.items():
                if judet in adresa_lower:
                    return anaf
        
        return "NEDETERMINAT"
    
    def add_single_record_to_excel(self, txt_file_path):
        """Adauga o singura inregistrare direct in fișierul Excel existent"""
        try:
            # Extragem datele din fișierul .txt
            data = self.extract_data_from_txt(txt_file_path)
            if not data:
                return False
            
            # Verificam daca fișierul Excel exista
            if os.path.exists(self.excel_file_path):
                # Citim Excel-ul existent
                try:
                    df_existing = pd.read_excel(self.excel_file_path, sheet_name='Date_Persoane')
                    
                    # Verificam daca aceasta inregistrare exista deja (pe baza caii fișierului)
                    if 'Cale_Fisier' in df_existing.columns:
                        relative_path_no_ext = os.path.splitext(os.path.relpath(txt_file_path, self.output_folder))[0]
                        if relative_path_no_ext in df_existing['Cale_Fisier'].values:
                            print(f"📋 Înregistrarea pentru {data['Nume']} {data['Prenume']} exista deja in Excel")
                            return True  # Consideram ca este ok, nu este o eroare
                            
                except:
                    # Daca nu poate citi, cream unul nou
                    df_existing = pd.DataFrame()
            else:
                # Cream un DataFrame gol
                df_existing = pd.DataFrame()
            
            # Cream DataFrame cu noua inregistrare
            df_new = pd.DataFrame([data])
            
            # Combinam cu datele existente
            if not df_existing.empty:
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            else:
                df_combined = df_new
            
            # Ne asiguram ca CNP-ul și Telefonul sunt tratate ca string pentru a evita formatarea automata
            if 'CNP' in df_combined.columns:
                df_combined['CNP'] = df_combined['CNP'].astype(str)
            if 'Telefon' in df_combined.columns:
                df_combined['Telefon'] = df_combined['Telefon'].astype(str)
            
            # Add a new column 'Validitate_CNP' next to 'CNP'
            if 'CNP' in df_combined.columns:
                df_combined['Validitate_CNP'] = df_combined['CNP'].apply(lambda cnp: 'DA' if self.validate_cnp(cnp)[0] else 'NU')
            
            # Reordonam coloanele pentru a respecta ordinea ceruta
            preferred_columns = [
                'Nume', 'Initiala_Tatalui', 'Prenume', 'CNP', 'Validitate_CNP', 'Adresa', 'ANAF_Apartin', 'Telefon', 'Email', '2_Ani'
            ]
            
            # Adaugam coloanele suplimentare la sfârșit pentru referinta
            additional_columns = [
                'Nume_Fisier', 'Folder_Sursa', 'Cale_Fisier', 'Data_Procesare', 'Text_Complet'
            ]
            
            # Reordonam coloanele existente
            existing_main_columns = [col for col in preferred_columns if col in df_combined.columns]
            existing_additional_columns = [col for col in additional_columns if col in df_combined.columns]
            other_columns = [col for col in df_combined.columns if col not in preferred_columns + additional_columns]
            
            final_columns = existing_main_columns + existing_additional_columns + other_columns
            
            df_combined = df_combined[final_columns]
            
            # Salvam in Excel cu formatare
            with pd.ExcelWriter(self.excel_file_path, engine='openpyxl') as writer:
                df_combined.to_excel(writer, sheet_name='Date_Persoane', index=False)
                
                # Obtinem worksheet-ul pentru formatare
                worksheet = writer.sheets['Date_Persoane']
                
                # Formatam coloanele CNP și Telefon ca text pentru a evita formatarea automata
                from openpyxl.styles import NamedStyle
                from openpyxl.utils import get_column_letter
                
                # Gasim coloanele CNP și Telefon
                columns_to_format = {'CNP': None, 'Telefon': None}
                for idx, col_name in enumerate(df_combined.columns, 1):
                    if col_name in columns_to_format:
                        columns_to_format[col_name] = idx
                
                print(f"Debug: Coloane de formatat gasite in add_single_record: {columns_to_format}")
                
                # Formatam coloanele gasite
                for col_name, col_index in columns_to_format.items():
                    if col_index:
                        col_letter = get_column_letter(col_index)
                        # Formatam intreaga coloana ca text (inclusiv header-ul)
                        for row in range(1, len(df_combined) + 2):  # De la header (1) pâna la ultimul rând
                            cell = worksheet[f'{col_letter}{row}']
                            cell.number_format = '@'  # Format text
                            
                            # Pentru datele din rândurile de continut (nu header)
                            if row > 1 and cell.value is not None:
                                # Convertim totul la string, indiferent de continut
                                original_value = str(cell.value).strip()
                                
                                # Pentru CNP, daca este in notatie știintifica, il convertim
                                if col_name == 'CNP' and ('E+' in original_value or 'e+' in original_value or '.' in original_value):
                                    try:
                                        # Încercam sa convertim din notatie știintifica
                                        cell.value = str(int(float(original_value)))
                                    except:
                                        cell.value = original_value
                                # Pentru telefon, eliminam .0 daca exista, dar pastram zero-urile de la inceput
                                elif col_name == 'Telefon' and original_value.endswith('.0'):
                                    # Eliminam doar .0 de la sfârșit, pastrând zero-urile de la inceput
                                    cell.value = original_value[:-2]
                                else:
                                    # Pentru toate celelalte, il pastram ca string
                                    cell.value = original_value
                
                # Ajustam latimea coloanelor
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Maxim 50 caractere
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"📋 Adaugata in Excel: {data['Nume']} {data['Prenume']}")
            return True
            
        except Exception as e:
            print(f"Eroare la adaugarea inregistrarii in Excel: {e}")
            return False

    def validate_cnp(self, cnp):
        """Valideaza CNP-ul conform algoritmului oficial românesc"""
        if not cnp or not isinstance(cnp, str):
            return False, "CNP invalid - nu este string"
            
        # Eliminam spatiile și convertim la string
        cnp = str(cnp).strip().replace(" ", "")
        
        # Verificam lungimea
        if len(cnp) != 13:
            return False, f"CNP invalid - lungime {len(cnp)} in loc de 13"
            
        # Verificam ca sunt doar cifre
        if not cnp.isdigit():
            return False, "CNP invalid - contine caractere non-numerice"
        
        # Verificam prima cifra (sexul și secolul)
        if cnp[0] not in '123456789':
            return False, f"CNP invalid - prima cifra {cnp[0]} nu este valida"
            
        # Verificam anul
        try:
            an = int(cnp[1:3])
            luna = int(cnp[3:5])
            zi = int(cnp[5:7])
            
            # Determinam secolul pe baza primei cifre
            if cnp[0] in '12':
                an += 1900
            elif cnp[0] in '34':
                an += 1800
            elif cnp[0] in '56':
                an += 2000
            elif cnp[0] in '78':
                an += 1800  # Rezidenti straini
            elif cnp[0] == '9':
                an += 2000  # Straini cu permis de ședere
                
            # Verificam validitatea datei
            if luna < 1 or luna > 12:
                return False, f"CNP invalid - luna {luna} nu este valida"
                
            if zi < 1 or zi > 31:
                return False, f"CNP invalid - ziua {zi} nu este valida"
                
            # Verificare simplificata pentru zile in functie de luna
            if luna in [4, 6, 9, 11] and zi > 30:
                return False, f"CNP invalid - luna {luna} nu poate avea {zi} zile"
                
            if luna == 2 and zi > 29:
                return False, f"CNP invalid - februarie nu poate avea {zi} zile"
                
        except ValueError:
            return False, "CNP invalid - data de naștere nu poate fi interpretata"
        
        # Verificam cifra de control (ultima cifra)
        multiplicatori = [2, 7, 9, 1, 4, 6, 3, 5, 8, 2, 7, 9]
        suma = sum(int(cnp[i]) * multiplicatori[i] for i in range(12))
        rest = suma % 11
        cifra_control = 1 if rest == 10 else rest
        
        if int(cnp[12]) != cifra_control:
            return False, f"CNP invalid - cifra de control {cnp[12]} nu corespunde cu calculul {cifra_control}"
            
        return True, "CNP valid"
    
    def _extract_person_data_from_txt(self, txt_file_path):
        """Extrage și convertește datele din txt pentru search index"""
        try:
            # Folosim funcția existentă pentru a extrage datele
            excel_data = self.extract_data_from_txt(txt_file_path)
            if not excel_data:
                return None
            
            # Convertim în formatul necesar pentru SearchManager
            search_data = {
                'nume': excel_data.get('Nume', ''),
                'prenume': excel_data.get('Prenume', ''),
                'cnp': excel_data.get('CNP', ''),
                'adresa': excel_data.get('Adresa', ''),
                'telefon': excel_data.get('Telefon', ''),
                'email': excel_data.get('Email', ''),
                'judet': self._extract_judet_from_address(excel_data.get('Adresa', '')),
                'localitate': self._extract_localitate_from_address(excel_data.get('Adresa', '')),
                'processing_date': excel_data.get('Data_Procesare', ''),
                'file_path': txt_file_path,
                'anaf_sector': excel_data.get('ANAF_Apartin', ''),
                'doiani': excel_data.get('2_Ani', '')
            }
            
            return search_data
            
        except Exception as e:
            print(f"Eroare la extragerea datelor pentru search: {e}")
            return None
    
    def _extract_judet_from_address(self, adresa):
        """Extrage județul din adresă"""
        if not adresa:
            return ''
        
        # Caută pattern-ul "JUD. numele_judetului"
        import re
        match = re.search(r'JUD\.?\s*([A-Z][a-zA-Z\s]+)', adresa, re.IGNORECASE)
        if match:
            judet = match.group(1).strip()
            # Curăță sfârșitul (elimină caracterele care nu sunt litere)
            judet = re.sub(r'[^a-zA-Z\s].*$', '', judet).strip()
            return judet
        return ''
    
    def _extract_localitate_from_address(self, adresa):
        """Extrage localitatea din adresă"""
        if not adresa:
            return ''
        
        # Caută pattern-ul "LOC. numele_localitatii"
        import re
        match = re.search(r'LOC\.?\s*([A-Z][a-zA-Z\s]+)', adresa, re.IGNORECASE)
        if match:
            localitate = match.group(1).strip()
            # Curăță sfârșitul până la următorul cuvânt cheie
            localitate = re.sub(r'[^a-zA-Z\s].*$', '', localitate).strip()
            return localitate
        return ''

    def detect_duplicate_entries(self):
        """Detecteaza duplicate in Excel pe baza CNP-ului și returneaza un raport"""
        if not os.path.exists(self.excel_file_path):
            return {
                'total_records': 0,
                'duplicates_found': 0,
                'duplicate_groups': [],
                'message': 'Fișierul Excel nu exista'
            }
            
        try:
            df = pd.read_excel(self.excel_file_path, sheet_name='Date_Persoane')
            
            if 'CNP' not in df.columns:
                return {
                    'total_records': len(df),
                    'duplicates_found': 0,
                    'duplicate_groups': [],
                    'message': 'Coloana CNP nu exista in Excel'
                }
            
            # Filtram CNP-urile valide (nu goale)
            df_valid_cnp = df[df['CNP'].notna() & (df['CNP'] != '') & (df['CNP'] != 'nan')]
            
            # Gasim duplicatele
            cnp_counts = df_valid_cnp['CNP'].value_counts()
            duplicates = cnp_counts[cnp_counts > 1]
            
            duplicate_groups = []
            for cnp, count in duplicates.items():
                duplicate_rows = df_valid_cnp[df_valid_cnp['CNP'] == cnp]
                group_info = {
                    'cnp': cnp,
                    'count': count,
                    'records': []
                }
                
                for _, row in duplicate_rows.iterrows():
                    record_info = {
                        'nume': row.get('Nume', ''),
                        'prenume': row.get('Prenume', ''),
                        'fisier': row.get('Nume_Fisier', ''),
                        'data_procesare': row.get('Data_Procesare', '')
                    }
                    group_info['records'].append(record_info)
                    
                duplicate_groups.append(group_info)
            
            return {
                'total_records': len(df),
                'valid_cnp_records': len(df_valid_cnp),
                'duplicates_found': len(duplicates),
                'total_duplicate_records': sum(duplicates.values),
                'duplicate_groups': duplicate_groups,
                'message': f'Gasite {len(duplicates)} CNP-uri duplicate in {sum(duplicates.values)} inregistrari'
            }
            
        except Exception as e:
            return {
                'total_records': 0,
                'duplicates_found': 0,
                'duplicate_groups': [],
                'message': f'Eroare la detectarea duplicatelor: {str(e)}'
            }
    
    def export_to_csv(self, csv_file_path=None):
        """Export in CSV pentru compatibilitate - salvează în folderul de output"""
        if not os.path.exists(self.excel_file_path):
            print("Fișierul Excel nu exista pentru export CSV")
            return False
            
        if csv_file_path is None:
            # Salvează în folderul de output, nu în același loc cu Excel-ul
            csv_filename = "Date_Persoane_OCR.csv"
            csv_file_path = os.path.join(self.output_folder, csv_filename)
            
        try:
            # Citim datele din Excel
            df = pd.read_excel(self.excel_file_path, sheet_name='Date_Persoane')
            
            # Exportam in CSV cu encoding UTF-8 BOM pentru Excel
            df.to_csv(csv_file_path, index=False, encoding='utf-8-sig', sep=';')
            
            print(f"✅ Export CSV realizat cu succes: {csv_file_path}")
            print(f"📊 Exportate {len(df)} inregistrari")
            return csv_file_path
            
        except Exception as e:
            print(f"❌ Eroare la exportul CSV: {e}")
            return False
    
    def export_to_pdf_report(self, pdf_file_path=None):
        """Genereaza raport PDF cu statistici, grafice și watermark pe fiecare pagină"""
        try:
            # Verificam daca matplotlib și reportlab sunt disponibile
            try:
                import matplotlib.pyplot as plt
                import matplotlib
                matplotlib.use('Agg')  # Backend non-interactiv
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.pdfgen import canvas as pdfcanvas
                from reportlab.lib.utils import ImageReader
            except ImportError:
                print("❌ Pentru export PDF este nevoie de: pip install matplotlib reportlab")
                return False
            
            if not os.path.exists(self.excel_file_path):
                print("Fișierul Excel nu exista pentru export PDF")
                return False
                
            if pdf_file_path is None:
                # Salvează în folderul de output cu nume sugestiv
                pdf_filename = "Raport_OCR_F230.pdf"
                pdf_file_path = os.path.join(self.output_folder, pdf_filename)
                
            # Citim datele
            df = pd.read_excel(self.excel_file_path, sheet_name='Date_Persoane')
            
            # Cream documentul PDF
            story = []
            styles = getSampleStyleSheet()
            
            # Titlu
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph("Raport OCR F230 - Analiza Datelor", title_style))
            story.append(Spacer(1, 20))
            
            # Statistici generale
            total_records = len(df)
            valid_cnp = len(df[df['CNP'].notna() & (df['CNP'] != '') & (df['CNP'] != 'nan')])
            
            # Calculez statistici pentru "2 ani"
            doiani_da = len(df[df['2_Ani'].str.lower().str.contains('da', na=False)])
            doiani_nu = total_records - doiani_da
            
            # Statistici ANAF
            anaf_stats = df['ANAF_Apartin'].value_counts()
            
            stats_text = f"""
            <b>Statistici Generale:</b><br/>
            • Total inregistrari procesate: {total_records}<br/>
            • Înregistrari cu CNP valid: {valid_cnp}<br/>
            • Persoane donatoare pe 2 ani: {doiani_da}<br/>
            • Persoane donatoare pe 1 an: {doiani_nu}<br/>
            • Data generare raport: {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
            """
            
            story.append(Paragraph(stats_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Distributia pe ANAF
            story.append(Paragraph("<b>Distributia pe ANAF:</b>", styles['Heading2']))
            
            anaf_data = [['ANAF', 'Numar persoane', 'Procent']]
            for anaf, count in anaf_stats.head(10).items():  # Top 10
                procent = (count / total_records) * 100
                anaf_data.append([anaf, str(count), f"{procent:.1f}%"])
                
            anaf_table = Table(anaf_data)
            anaf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(anaf_table)
            story.append(Spacer(1, 20))
            
            # Detectare duplicate
            duplicate_info = self.detect_duplicate_entries()
            duplicate_text = f"""
            <b>Verificare Duplicate:</b><br/>
            • CNP-uri duplicate gasite: {duplicate_info['duplicates_found']}<br/>
            • Total inregistrari duplicate: {duplicate_info.get('total_duplicate_records', 0)}<br/>
            """
            
            if duplicate_info['duplicates_found'] > 0:
                duplicate_text += "<b>⚠️ Atentie: Exista CNP-uri duplicate in baza de date!</b><br/>"
                
            story.append(Paragraph(duplicate_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Generam și salvam graficul
            plt.figure(figsize=(10, 6))
            
            # Grafic cu distributia 1 an vs 2 ani
            plt.subplot(1, 2, 1)
            labels = ['1 an', '2 ani']
            sizes = [doiani_nu, doiani_da]
            colors_pie = ['#ff9999', '#66b3ff']
            plt.pie(sizes, labels=labels, colors=colors_pie, autopct='%1.1f%%', startangle=90)
            plt.title('Distributia perioadei de donatie')
            
            # Grafic cu top 5 ANAF
            plt.subplot(1, 2, 2)
            top_anaf = anaf_stats.head(5)
            plt.bar(range(len(top_anaf)), top_anaf.values, color='skyblue')
            plt.title('Top 5 ANAF dupa numarul de persoane')
            plt.xlabel('ANAF')
            plt.ylabel('Numarul de persoane')
            plt.xticks(range(len(top_anaf)), [anaf[:15] + '...' if len(anaf) > 15 else anaf for anaf in top_anaf.index], rotation=45, ha='right')
            
            plt.tight_layout()
            
            # Salvam graficul ca imagine temporara
            chart_path = pdf_file_path.replace('.pdf', '_chart.png')
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            # Adaugam graficul in PDF
            story.append(Spacer(1, 20))
            story.append(Paragraph("<b>Grafice & Analize:</b>", styles['Heading2']))
            story.append(Spacer(1, 10))
            story.append(Image(chart_path, width=6*inch, height=3.6*inch))
            
            # Funcție pentru watermark pe fiecare pagină
            def add_watermark(canvas, doc):
                canvas.setTitle("Raport OCR F230")
                watermark_path = os.path.join(os.path.dirname(__file__), '../../Assets/favicon.png')
                watermark_path = os.path.abspath(watermark_path)
                try:
                    img = ImageReader(watermark_path)
                    page_width, page_height = A4
                    # Dimensiune watermark (ajustabilă)
                    wm_width = page_width * 0.4
                    wm_height = page_width * 0.4
                    x = (page_width - wm_width) / 2
                    y = (page_height - wm_height) / 2
                    canvas.saveState()
                    canvas.setFillAlpha(0.15)
                    canvas.drawImage(img, x, y, width=wm_width, height=wm_height, mask='auto', preserveAspectRatio=True)
                    canvas.restoreState()
                except Exception as e:
                    print(f"[Watermark] Eroare la adăugarea watermarkului: {e}")
            
            # Construim PDF-ul cu watermark pe fiecare pagină
            doc = SimpleDocTemplate(pdf_file_path, pagesize=A4)
            doc.build(story, onFirstPage=add_watermark, onLaterPages=add_watermark)
            
            # Ștergem imaginea temporara
            try:
                os.remove(chart_path)
            except:
                pass
                
            print(f"✅ Raport PDF generat cu succes: {pdf_file_path}")
            return pdf_file_path
            
        except Exception as e:
            print(f"❌ Eroare la generarea raportului PDF: {e}")
            return False

    def validate_and_clean_data(self):
        """Valideaza și curata datele din lista curenta"""
        if not self.data_list:
            return {'valid': 0, 'invalid': 0, 'errors': []}
            
        valid_count = 0
        invalid_count = 0
        validation_errors = []
        
        for i, data in enumerate(self.data_list):
            errors_for_record = []
            
            # Validare CNP
            if data.get('CNP'):
                is_valid, message = self.validate_cnp(data['CNP'])
                print('=='*60)
                print(data['CNP'], is_valid, message)
                print('=='*60)
                if not is_valid:
                    errors_for_record.append(f"CNP: {message}")
                    invalid_count += 1
                else:
                    valid_count += 1
            else:
                errors_for_record.append("CNP: Lipsește")
                invalid_count += 1
                
            # Validare telefon (format românesc basic)
            if data.get('Telefon'):
                telefon = str(data['Telefon']).strip()
                if telefon and not re.match(r'^(0[0-9]{9}|(\+40[0-9]{9}))$', telefon.replace(' ', '')):
                    errors_for_record.append(f"Telefon: Format invalid '{telefon}'")
                    
            # Validare email basic
            if data.get('Email'):
                email = str(data['Email']).strip()
                if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                    errors_for_record.append(f"Email: Format invalid '{email}'")
                    
            # Validare nume (nu poate fi gol)
            if not data.get('Nume') or not data.get('Nume').strip():
                errors_for_record.append("Nume: Lipsește sau este gol")
                
            if errors_for_record:
                validation_errors.append({
                    'record_index': i,
                    'nume': data.get('Nume', 'N/A'),
                    'prenume': data.get('Prenume', 'N/A'),
                    'fisier': data.get('Nume_Fisier', 'N/A'),
                    'errors': errors_for_record
                })
                
        return {
            'valid': valid_count,
            'invalid': invalid_count,
            'total_records': len(self.data_list),
            'validation_errors': validation_errors
        }

    def generate_validation_report(self):
        """Genereaza un raport de validare pentru datele curente"""
        validation_result = self.validate_and_clean_data()
        duplicate_result = self.detect_duplicate_entries()
        
        print("\n" + "="*60)
        print("📋 RAPORT DE VALIDARE DATEELOR OCR F230")
        print("="*60)
        
        print(f"\n📊 STATISTICI GENERALE:")
        print(f"   • Total inregistrari: {validation_result['total_records']}")
        print(f"   • CNP-uri valide: {validation_result['valid']}")
        print(f"   • CNP-uri invalide: {validation_result['invalid']}")
        print(f"   • CNP-uri duplicate: {duplicate_result['duplicates_found']}")
        
        if validation_result['validation_errors']:
            print(f"\n⚠️  ERORI DE VALIDARE ({len(validation_result['validation_errors'])}):")
            for error in validation_result['validation_errors'][:10]:  # Afișam doar primele 10
                print(f"   📄 {error['fisier']} - {error['nume']} {error['prenume']}:")
                for err in error['errors']:
                    print(f"      ❌ {err}")
                    
            if len(validation_result['validation_errors']) > 10:
                print(f"   ... și inca {len(validation_result['validation_errors']) - 10} erori")
                
        if duplicate_result['duplicate_groups']:
            print(f"\n🔄 CNP-URI DUPLICATE ({len(duplicate_result['duplicate_groups'])}):")
            for dup in duplicate_result['duplicate_groups'][:5]:  # Afișam doar primele 5
                print(f"   🆔 CNP: {dup['cnp']} ({dup['count']} inregistrari)")
                for record in dup['records']:
                    print(f"      • {record['nume']} {record['prenume']} - {record['fisier']}")
                    
        print("\n" + "="*60)
        
        return validation_result, duplicate_result

def create_excel_summary(output_folder):
    """Functie principala pentru crearea rezumatului Excel, PDF și CSV"""
    try:
        excel_manager = ExcelManager(output_folder)
        success = excel_manager.process_all_txt_files(output_folder)
        
        if success:
            print(f"✅ Excel creat cu succes!")
            
            # Generăm automat și CSV-ul
            print(f"📋 Generez CSV...")
            csv_path = excel_manager.export_to_csv()
            
            # Generăm automat și PDF-ul
            print(f"📄 Generez raportul PDF...")
            pdf_path = excel_manager.export_to_pdf_report()
            
            print(f"🎯 Fișiere generate:")
            print(f"   📊 Excel: Date_Persoane_OCR.xlsx")
            if csv_path:
                print(f"   📋 CSV: Date_Persoane_OCR.csv")
            if pdf_path:
                print(f"   📄 PDF: Raport_OCR_F230.pdf")
            
            return excel_manager.excel_file_path
        else:
            return None
            
    except Exception as e:
        print(f"Eroare la crearea rezumatului Excel: {e}")
        return None

    def _extract_person_data_from_txt(self, txt_file_path):
        """Extrage și convertește datele din txt pentru search index"""
        try:
            # Folosim funcția existentă pentru a extrage datele
            excel_data = self.extract_data_from_txt(txt_file_path)
            if not excel_data:
                return None
            
            # Convertim în formatul necesar pentru SearchManager
            search_data = {
                'nume': excel_data.get('Nume', ''),
                'prenume': excel_data.get('Prenume', ''),
                'cnp': excel_data.get('CNP', ''),
                'adresa': excel_data.get('Adresa', ''),
                'telefon': excel_data.get('Telefon', ''),
                'email': excel_data.get('Email', ''),
                'judet': self._extract_judet_from_address(excel_data.get('Adresa', '')),
                'localitate': self._extract_localitate_from_address(excel_data.get('Adresa', '')),
                'processing_date': excel_data.get('Data_Procesare', ''),
                'file_path': txt_file_path,
                'anaf_sector': excel_data.get('ANAF_Apartin', ''),
                'doiani': excel_data.get('2_Ani', '')
            }
            
            return search_data
            
        except Exception as e:
            print(f"Eroare la extragerea datelor pentru search: {e}")
            return None
    
    def _extract_judet_from_address(self, adresa):
        """Extrage județul din adresă"""
        if not adresa:
            return ''
        
        # Caută pattern-ul "JUD. numele_judetului"
        import re
        match = re.search(r'JUD\.?\s*([A-Z][a-zA-Z\s]+)', adresa, re.IGNORECASE)
        if match:
            judet = match.group(1).strip()
            # Curăță sfârșitul (elimină caracterele care nu sunt litere)
            judet = re.sub(r'[^a-zA-Z\s].*$', '', judet).strip()
            return judet
        return ''
    
    def _extract_localitate_from_address(self, adresa):
        """Extrage localitatea din adresă"""
        if not adresa:
            return ''
        
        # Caută pattern-ul "LOC. numele_localitatii"
        import re
        match = re.search(r'LOC\.?\s*([A-Z][a-zA-Z\s]+)', adresa, re.IGNORECASE)
        if match:
            localitate = match.group(1).strip()
            # Curăță sfârșitul până la următorul cuvânt cheie
            localitate = re.sub(r'[^a-zA-Z\s].*$', '', localitate).strip()
            return localitate
        return ''

def add_single_person_to_excel(output_folder, txt_file_path):
    """Functie pentru a adauga o singura persoana in Excel și actualiza CSV/PDF"""
    try:
        excel_manager = ExcelManager(output_folder)
        success = excel_manager.add_single_record_to_excel(txt_file_path)
        
        if success:
            # Actualizăm și CSV-ul și PDF-ul după adăugarea unei noi persoane
            print(f"📋 Actualizez CSV...")
            excel_manager.export_to_csv()
            
            print(f"📄 Actualizez raportul PDF...")
            excel_manager.export_to_pdf_report()
            
            # 🔍 ADAUGĂ PERSOANA ÎN SEARCH INDEX
            try:
                print(f"🔍 Adaug în search index...")
                from src.search.search_manager import SearchManager
                search_manager = SearchManager(output_folder)
                
                # Extrag datele din fișierul txt pentru search
                person_data = excel_manager._extract_person_data_from_txt(txt_file_path)
                if person_data:
                    search_id = search_manager.add_person_to_index(person_data)
                    print(f"✅ Adăugat în search index cu ID: {search_id}")
                else:
                    print("⚠️ Nu s-au putut extrage datele pentru search")
                    
            except Exception as search_error:
                print(f"⚠️ Eroare la adăugarea în search index: {search_error}")
                # Nu opresc procesarea pentru că Excel a fost deja actualizat
            
        return success
    except Exception as e:
        print(f"Eroare la adaugarea persoanei in Excel: {e}")
        return False
